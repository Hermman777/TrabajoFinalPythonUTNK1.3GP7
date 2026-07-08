"""
control de asistencia
Filtro de socios por tipo de membresia
El administrador va cargando por teclado los datos de usuario y el programa los almacena en un archivo excel en el primer uso, si el archivo no existe, el programa lo crea vacio. En los usos siguientes, el programa lee ese archivo para recuperar todo lo cargado anteriormente y permite seguir administrandolo
"""

import os
from openpyxl import Workbook, load_workbook

NOMBRE_ARCHIVO = "gimnasio_datos.xlsx"
RUTA_EXCEL = os.path.join(os.path.dirname(os.path.abspath(__file__)), NOMBRE_ARCHIVO)

HOJA_SOCIOS = "Socios"
HOJA_ACTIVIDADES = "Actividades"
HOJA_INSCRIPCIONES = "Inscripciones"
HOJA_ASISTENCIAS = "Asistencias"

ENCABEZADOS_SOCIOS = ["DNI", "Nombre", "Membresia", "CostoCuota", "Activo", "CuotasPagadas"]
ENCABEZADOS_ACTIVIDADES = ["Nombre", "Cupo"]
ENCABEZADOS_INSCRIPCIONES = ["DNI", "Actividad"]
ENCABEZADOS_ASISTENCIAS = ["DNI", "Actividad"]

TIPOS_MEMBRESIA = {
    "1": {"nombre": "Básica", "costo": 5000},
    "2": {"nombre": "Full",   "costo": 8000},
    "3": {"nombre": "Estudiante (promo)", "costo": 4000},
}

socios = []                 
actividades = []             
registro_asistencias = []    
asistencias_por_socio = {}       
asistencias_por_actividad = {}   
total_recaudado = 0.0     
contador_asistencias = 0  


def crear_archivo_excel():
    """crea el archivo Excel la primera vez que se usa el sistema"""
    libro = Workbook()
    hoja = libro.active
    hoja.title = HOJA_SOCIOS
    hoja.append(ENCABEZADOS_SOCIOS)
    hoja = libro.create_sheet(HOJA_ACTIVIDADES)
    hoja.append(ENCABEZADOS_ACTIVIDADES)
    hoja = libro.create_sheet(HOJA_INSCRIPCIONES)
    hoja.append(ENCABEZADOS_INSCRIPCIONES)
    hoja = libro.create_sheet(HOJA_ASISTENCIAS)
    hoja.append(ENCABEZADOS_ASISTENCIAS)
    libro.save(RUTA_EXCEL)
    print(f"se creo el archivo '{NOMBRE_ARCHIVO}' para guardar los datos")


def cargar_datos_desde_excel():
    """lee el excel existente y reconstruye todas las estructuras en memoria"""
    global total_recaudado, contador_asistencias
    libro = load_workbook(RUTA_EXCEL)
    hoja = libro[HOJA_SOCIOS]
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        dni, nombre, membresia, costo_cuota, activo, cuotas_pagadas = fila
        socios.append({
            "dni": str(dni),
            "nombre": nombre,
            "membresia": membresia,
            "costo_cuota": float(costo_cuota) if costo_cuota is not None else 0.0,
            "activo": bool(activo),
            "cuotas_pagadas": int(cuotas_pagadas) if cuotas_pagadas is not None else 0,
        })
        total_recaudado += socios[-1]["costo_cuota"] * socios[-1]["cuotas_pagadas"]
    hoja = libro[HOJA_ACTIVIDADES]
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        nombre, cupo = fila
        actividades.append({"nombre": nombre, "cupo": int(cupo), "inscriptos": []})
        asistencias_por_actividad[nombre] = 0
    hoja = libro[HOJA_INSCRIPCIONES]
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        dni, nombre_actividad = fila
        actividad = buscar_actividad_por_nombre(nombre_actividad)
        if actividad is not None:
            actividad["inscriptos"].append(str(dni))
    hoja = libro[HOJA_ASISTENCIAS]
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[0] is None:
            continue
        dni, nombre_actividad = fila
        dni = str(dni)
        registro_asistencias.append({"dni": dni, "actividad": nombre_actividad})
        asistencias_por_socio[dni] = asistencias_por_socio.get(dni, 0) + 1
        asistencias_por_actividad[nombre_actividad] = asistencias_por_actividad.get(nombre_actividad, 0) + 1
        contador_asistencias += 1

    print(f"Datos cargados desde '{NOMBRE_ARCHIVO}': {len(socios)} socio "
          f"{len(actividades)} actividad, {contador_asistencias} asistencia registradas")

def guardar_datos():
    """Vuelca todo el estado actual en memoria hacia el archivo excel"""
    try:
        libro = Workbook()
        hoja = libro.active
        hoja.title = HOJA_SOCIOS
        hoja.append(ENCABEZADOS_SOCIOS)
        for socio in socios:
            hoja.append([socio["dni"], socio["nombre"], socio["membresia"],
                         socio["costo_cuota"], socio["activo"], socio["cuotas_pagadas"]])
        hoja = libro.create_sheet(HOJA_ACTIVIDADES)
        hoja.append(ENCABEZADOS_ACTIVIDADES)
        for actividad in actividades:
            hoja.append([actividad["nombre"], actividad["cupo"]])
        hoja = libro.create_sheet(HOJA_INSCRIPCIONES)
        hoja.append(ENCABEZADOS_INSCRIPCIONES)
        for actividad in actividades:
            for dni in actividad["inscriptos"]:
                hoja.append([dni, actividad["nombre"]])
        hoja = libro.create_sheet(HOJA_ASISTENCIAS)
        hoja.append(ENCABEZADOS_ASISTENCIAS)
        for registro in registro_asistencias:
            hoja.append([registro["dni"], registro["actividad"]])
        libro.save(RUTA_EXCEL)
    except PermissionError:
        print(f"Error no se pudo guardar '{NOMBRE_ARCHIVO}' "
              f"Cerra el archivo si lo tenes abierto y intenta denuevo")
    except Exception as error:
        print(f"Error inesperado al guardar los datos: {error}")

def inicializar_sistema():
    """al arrancar el programa crea el Excel si es la primera vez o lo carga si ya existe"""
    try:
        if not os.path.exists(RUTA_EXCEL):
            crear_archivo_excel()
        else:
            cargar_datos_desde_excel()
    except Exception as error:
        print(f"Error al inicializar la base de datos{error} "
              f"Se continuara con datos vacios")

def leer_texto_no_vacio(mensaje):
    """Pide un texto por consola hasta que no este vacio"""
    while True:
        try:
            texto = input(mensaje).strip()
            if texto == "":
                raise ValueError("El dato no puede estar vacio")
            return texto
        except ValueError as error:
            print(f"Error {error} intente otravez")

def leer_dni():
    """Valida que el DNI ingresado sea numerico"""
    while True:
        try:
            dni = input("ingrese DNI del socio: ").strip()
            if not dni.isdigit():
                raise ValueError("El DNI debe contener solo numeross")
            if not (6 <= len(dni) <= 9):
                raise ValueError("El DNI debe tener entre y 9 digitos")
            return dni
        except ValueError as error:
            print(f"Error {error} intente otra vez")


def leer_entero_positivo(mensaje):
    """Pide un numero entero positivo manejando errores de conversion"""
    while True:
        try:
            valor = int(input(mensaje))
            if valor <= 0:
                raise ValueError("El numero debe ser mayor a cero")
            return valor
        except ValueError as error:
            print(f"Error dato invalido {error} intente nuevamente")


def leer_opcion(mensaje, opciones_validas):
    """Pide una opcion de menu y valida que exista dentro de las opciones permitidas"""
    while True:
        opcion = input(mensaje).strip()
        if opcion in opciones_validas:
            return opcion
        print(f"Opcion invalida las opciones validas son: {', '.join(opciones_validas)}")


def buscar_socio_por_dni(dni):
    """Devuelve el diccionario del socio si existe o ninguna si no fue encontrado"""
    for socio in socios:
        if socio["dni"] == dni:
            return socio
    return None


def buscar_actividad_por_nombre(nombre):
    """Devuelve el diccionario de la actividad si existe o ninguna si no fue encontrada"""
    for actividad in actividades:
        if actividad["nombre"].lower() == nombre.lower():
            return actividad
    return None

def registrar_socio():
    """Da de alta un nuevo socio validando que el DNI no este repetido"""
    print("\nAlta de socio")
    dni = leer_dni()

    if buscar_socio_por_dni(dni) is not None:
        print("Error ya existe un socio registrado con ese DN")
        return

    nombre = leer_texto_no_vacio("ingrese nombre y apellido: ")

    print("Tipos de membresia disponibles:")
    for clave, datos in TIPOS_MEMBRESIA.items():
        print(f"  {clave}. {datos['nombre']} (${datos['costo']})")
    opcion_membresia = leer_opcion("Elija tipo de membresia: ", TIPOS_MEMBRESIA.keys())

    nuevo_socio = {
        "dni": dni,
        "nombre": nombre,
        "membresia": TIPOS_MEMBRESIA[opcion_membresia]["nombre"],
        "costo_cuota": TIPOS_MEMBRESIA[opcion_membresia]["costo"],
        "activo": True,
        "cuotas_pagadas": 0,
    }
    socios.append(nuevo_socio)
    guardar_datos()
    print(f"Socio '{nombre}' registrado con exito y guardado en '{NOMBRE_ARCHIVO}'")


def dar_baja_socio():
    """Da de baja a un socio existente"""
    print("\nBaja de socio")
    if len(socios) == 0:
        print("No hay socios cargados todavia")
        return

    dni = leer_dni()
    socio = buscar_socio_por_dni(dni)

    if socio is None:
        print("Error no existe un socio con ese DNI")
    elif not socio["activo"]:
        print("Ese socio ya se encuentra inactivo")
    else:
        socio["activo"] = False
        guardar_datos()
        print(f"Socio '{socio['nombre']}' dado de baja correctamente")


def listar_socios():
    """Lista todos los socios y cuantos estan activos/inactivos"""
    print("\nListado de socios")
    if len(socios) == 0:
        print("No hay socios cargados")
        return
    contador_activos = 0
    contador_inactivos = 0
    for socio in socios:
        estado = "Activo" if socio["activo"] else "Inactivo"
        if socio["activo"]:
            contador_activos += 1
        else:
            contador_inactivos += 1
        print(f"  DNI: {socio['dni']} {socio['nombre']} "
              f"Membresia: {socio['membresia']} Estado: {estado} "
              f"Cuotas pagadas: {socio['cuotas_pagadas']}")

    print(f"\nTotal socios {len(socios)} Activos: {contador_activos} Inactivos: {contador_inactivos}")


def imprimir_socios(lista_filtrada):
    """Imprime una lista de socios ya filtrada junto con la cantidad de coincidencias"""
    if len(lista_filtrada) == 0:
        print("No se encontraron socios que cumplan ese filtro")
        return

    for socio in lista_filtrada:
        estado = "Activo" if socio["activo"] else "Inactivo"
        print(f"  DNI: {socio['dni']} {socio['nombre']} "
              f"Membresia: {socio['membresia']} Estado: {estado} "
              f"Cuotas pagadas: {socio['cuotas_pagadas']}")
    print(f"\nCoincidencias encontradas: {len(lista_filtrada)}")

def filtrar_por_membresia():
    """Filtra y muestra los socios que tienen un tipo de membresia determinado"""
    print("\nFiltrar socios por tipo de membresia")
    print("Tipos de membresa disponibles:")
    for clave, datos in TIPOS_MEMBRESIA.items():
        print(f"  {clave}. {datos['nombre']}")
    opcion = leer_opcion("Elija el tipo de membresia a filtrar: ", TIPOS_MEMBRESIA.keys())
    nombre_membresia = TIPOS_MEMBRESIA[opcion]["nombre"]
    resultado = [socio for socio in socios if socio["membresia"] == nombre_membresia]
    print(f"\nSocios con membresia '{nombre_membresia}':")
    imprimir_socios(resultado)

def filtrar_por_estado():
    """Filtra y muestra los socios activos o inactivos"""
    print("\nFiltrar socios por estado")
    print("1 Activos")
    print("2 Inactivos")
    opcion = leer_opcion("Elija una opcion: ", ["1", "2"])
    buscar_activo = (opcion == "1")
    resultado = [socio for socio in socios if socio["activo"] == buscar_activo]
    print(f"\nSocios {'activos' if buscar_activo else 'inactivos'}:")
    imprimir_socios(resultado)


def filtrar_por_actividad():
    """Filtra y muestra los socios inscriptos en una actividad determinada"""
    print("\nFiltrar socios por actividad")
    if len(actividades) == 0:
        print("No hay actividades cargadas todavia")
        return
    for actividad in actividades:
        print(f"  - {actividad['nombre']}")
    nombre_actividad = leer_texto_no_vacio("Ingrese el nombre de la actividad: ")
    actividad = buscar_actividad_por_nombre(nombre_actividad)
    if actividad is None:
        print("Error esa actividad no existe")
        return
    resultado = [socio for socio in socios if socio["dni"] in actividad["inscriptos"]]
    print(f"\nSocios inscriptos en '{actividad['nombre']}':")
    imprimir_socios(resultado)

def registrar_pago_cuota():
    """Registra el pago de una cuota y acumula el dinero recaudado"""
    global total_recaudado
    print("\nRegistrar pago de cuota")
    dni = leer_dni()
    socio = buscar_socio_por_dni(dni)

    if socio is None:
        print("Error no existe un socio con ese DNI")
        return
    if not socio["activo"]:
        print("Error el socio este inactivo, no puede pagar cuota")
        return

    socio["cuotas_pagadas"] += 1
    total_recaudado += socio["costo_cuota"]
    guardar_datos()
    print(f"Pago registrado {socio['nombre']} pago ${socio['costo_cuota']} "
          f"{socio['membresia']} Total de cuotas pagadas: {socio['cuotas_pagadas']}.")


def ver_estado_cuotas():
    """Muestra el estado de cuotas de cada socio y el total recaudado"""
    print("\nEstado de cuota")
    if len(socios) == 0:
        print("No hay socios cargados todavia")
        return
    for socio in socios:
        print(f"  {socio['nombre']} (DNI {socio['dni']})"
              f"Cuotas pagadas: {socio['cuotas_pagadas']} Membresia: {socio['membresia']}")
    print(f"\nTotal recaudado acumulado: ${total_recaudado:.2f}")

def alta_actividad():
    """Da de alta una nueva actividad con un cupo maximo de inscriptos"""
    print("\nAlta de actividad")
    nombre = leer_texto_no_vacio("Nombre de la actividad")
    if buscar_actividad_por_nombre(nombre) is not None:
        print("Error ya existe una actividad con ese nombre")
        return
    cupo = leer_entero_positivo("Cupo maximo de socios: ")
    actividades.append({"nombre": nombre, "cupo": cupo, "inscriptos": []})
    asistencias_por_actividad[nombre] = 0
    guardar_datos()
    print(f"Actividad '{nombre}' creada con cupo para {cupo} socios")

def inscribir_socio_actividad():
    """Inscribe a un socio activo en una actividad"""
    print("\nInscripcion a actividad")
    if len(actividades) == 0:
        print("No hay actividades cargadas ")
        return
    dni = leer_dni()
    socio = buscar_socio_por_dni(dni)
    if socio is None:
        print("Error no existe un socio con ese DNI")
        return
    if not socio["activo"]:
        print("Error el socio esta inactivo, no puede inscribirse")
        return
    print("Actividades disponibles:")
    for actividad in actividades:
        print(f"{actividad['nombre']} (inscriptos: {len(actividad['inscriptos'])}/{actividad['cupo']})")
    nombre_actividad = leer_texto_no_vacio("Ingrese el nombre de la actividad: ")
    actividad = buscar_actividad_por_nombre(nombre_actividad)
    if actividad is None:
        print("Error esa actividad no existe")
    elif dni in actividad["inscriptos"]:
        print("El socio ya esta inscripto en esa actividad")
    elif len(actividad["inscriptos"]) >= actividad["cupo"]:
        print("Error no hay cupo disponible en esta actividad")
    else:
        actividad["inscriptos"].append(dni)
        guardar_datos()
        print(f"{socio['nombre']} fue inscripto en '{actividad['nombre']}'")

def listar_actividades():
    """Lista todas las actividades con su ocupacion de cupo"""
    print("\nListado de actividades")
    if len(actividades) == 0:
        print("No hay actividades cargadas todavia")
        return
    for actividad in actividades:
        ocupacion = len(actividad["inscriptos"])
        print(f"  {actividad['nombre']} Ocupacion: {ocupacion}/{actividad['cupo']}")

def registrar_asistencia():
    """Registra la asistencia de un socio a una actividad en la que esta inscripto"""
    global contador_asistencias
    print("\nRegistrar asistencia")
    dni = leer_dni()
    socio = buscar_socio_por_dni(dni)
    if socio is None:
        print("Error no existe un socio con ese DNI")
        return
    if not socio["activo"]:
        print("Error el socio esta inactivo")
        return
    nombre_actividad = leer_texto_no_vacio("Ingrese la actividad a la que asiste: ")
    actividad = buscar_actividad_por_nombre(nombre_actividad)
    if actividad is None:
        print("Error esa actividad no existe")
        return
    if dni not in actividad["inscriptos"]:
        print("Error el socio no esta inscripto en esa actividad")
        return
    registro_asistencias.append({"dni": dni, "actividad": actividad["nombre"]})
    asistencias_por_socio[dni] = asistencias_por_socio.get(dni, 0) + 1
    asistencias_por_actividad[actividad["nombre"]] = asistencias_por_actividad.get(actividad["nombre"], 0) + 1
    contador_asistencias += 1
    guardar_datos()
    print(f"Asistencia registrada para {socio['nombre']} en '{actividad['nombre']}'")

def ver_estadisticas_asistencia():
    """Muestra estadisticas basicas de concurrencia"""
    print("\nEstadisticas de asistencia")
    if contador_asistencias == 0:
        print("no se registraron asistencias")
        return
    print(f"Total de asistencias registradas: {contador_asistencias}")
    print("\nAsistencias por actividad:")
    for actividad in actividades:
        nombre = actividad["nombre"]
        cantidad = asistencias_por_actividad.get(nombre, 0)
        inscriptos = len(actividad["inscriptos"])
        promedio = cantidad / inscriptos if inscriptos > 0 else 0
        print(f"  {nombre}: {cantidad} asistencias, Promedio por inscripto: {promedio:.2f}")

    print("\nAsistencias por socio:")
    for socio in socios:
        cantidad = asistencias_por_socio.get(socio["dni"], 0)
        print(f"  {socio['nombre']}: {cantidad} asistencias")

def menu_socios():
    while True:
        print("\ngestion de socios")
        print("1 registrar socio")
        print("2 dar de baja socio")
        print("3 listar todos los socios")
        print("4 filtrar socios (membresia/estado/actividad)")
        print("0 volver al menu principal")
        opcion = leer_opcion("Seleccione una opcion: ", ["0", "1", "2", "3", "4"])
        if opcion == "1":
            registrar_socio()
        elif opcion == "2":
            dar_baja_socio()
        elif opcion == "3":
            listar_socios()
        elif opcion == "4":
            menu_filtros()
        elif opcion == "0":
            break

def menu_filtros():
    while True:
        print("\nFiltrar socios")
        print("1 por tipo de membresia")
        print("2 por estado (activo/inactivo)")
        print("3 por actividad")
        print("0 volver")
        opcion = leer_opcion("Seleccione una opcion: ", ["0", "1", "2", "3"])
        if opcion == "1":
            filtrar_por_membresia()
        elif opcion == "2":
            filtrar_por_estado()
        elif opcion == "3":
            filtrar_por_actividad()
        elif opcion == "0":
            break

def menu_cuotas():
    while True:
        print("\ngestion de cuotas")
        print("1 registrar pago de cuota")
        print("2 ver estado de cuotas")
        print("0 volver al menu principal")
        opcion = leer_opcion("Seleccione una opcion: ", ["0", "1", "2"])

        if opcion == "1":
            registrar_pago_cuota()
        elif opcion == "2":
            ver_estado_cuotas()
        elif opcion == "0":
            break


def menu_actividades():
    while True:
        print("\ngestion de actividades")
        print("1 alta de actividad")
        print("2 inscribir socio a actividad")
        print("3 listar actividades")
        print("0 volver al menu principal")
        opcion = leer_opcion("Seleccione una opcion: ", ["0", "1", "2", "3"])
        if opcion == "1":
            alta_actividad()
        elif opcion == "2":
            inscribir_socio_actividad()
        elif opcion == "3":
            listar_actividades()
        elif opcion == "0":
            break

def menu_asistencia():
    while True:
        print("\ngestion de asistenci")
        print("1 registrar asistencia")
        print("2 ver estadisticas de asistencia")
        print("0 volver al menu principal")
        opcion = leer_opcion("Seleccione una opcion: ", ["0", "1", "2"])
        if opcion == "1":
            registrar_asistencia()
        elif opcion == "2":
            ver_estadisticas_asistencia()
        elif opcion == "0":
            break

def menu_principal():
    """Bucle principal del sistema Atrapa errores inesperados para que el programa no se cierre"""
    while True:
        print("gestion de gimnasio")
        print(f"(datos guardados en: {NOMBRE_ARCHIVO})")
        print("1 gestion de socios")
        print("2 gestion de cuotas")
        print("3 gestion de actividades")
        print("4 control de asistencia")
        print("0 salir")
        opcion = leer_opcion("Seleccione una opcion: ", ["0", "1", "2", "3", "4"])
        try:
            if opcion == "1":
                menu_socios()
            elif opcion == "2":
                menu_cuotas()
            elif opcion == "3":
                menu_actividades()
            elif opcion == "4":
                menu_asistencia()
            elif opcion == "0":
                print("\nSaliendo del sistema. ¡Hasta luego!")
                break
        except Exception as error:
            print(f"Ocurrio un error inesperado {error} Volviendo al menu")


if __name__ == "__main__":
    try:
        inicializar_sistema()
        menu_principal()
    except KeyboardInterrupt:
        print("\n\nPrograma interrumpido por el usuario")
    except EOFError:
        print("\n\nNo se pudo seguir leyendo datos de entrada, cerrando el programa")
